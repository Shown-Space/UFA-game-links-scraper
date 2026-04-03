#!/usr/bin/env python3
"""
UFA Game WatchUFA Link Finder
==============================
Parses the watchufa.tv sitemap to extract game stream URLs and matches
them to games in the database by team + date.

SETUP:
  pip install openpyxl python-dotenv

.env file needs:
  SUPABASE_URL=...
  SUPABASE_KEY=...

OUTPUT:
  ufa_watchufa_links.xlsx  — spreadsheet of all games with WatchUFA URLs
  ufa_watchufa_update.sql  — SQL to update the database
"""

import os
import re
import json
import urllib.request
from datetime import date, datetime, timedelta
from xml.etree import ElementTree

from dotenv import load_dotenv

load_dotenv()

# ─── CONFIGURATION ────────────────────────────────────────────────────────────

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

YEAR_FILTER   = 2025
SITEMAP_PATH  = "sitemap.xml"
OUTPUT_XLSX   = "ufa_watchufa_links.xlsx"
OUTPUT_SQL    = "ufa_watchufa_update.sql"

# ─── CITY SLUG → TEAM ID ──────────────────────────────────────────────────────

CITY_SLUG_TO_TEAM = {
    "atlanta":      "hustle",
    "austin":       "sol",
    "boston":       "glory",
    "carolina":     "flyers",
    "chicago":      "union",
    "colorado":     "apex",
    "dc":           "breeze",
    "detroit":      "mechanix",
    "houston":      "havoc",
    "indianapolis": "alleycats",
    "las-vegas":    "bighorns",
    "los-angeles":  "aviators",
    "madison":      "radicals",
    "minnesota":    "windchill",
    "montreal":     "royal",
    "new-york":     "empire",
    "oakland":      "spiders",
    "oregon":       "steel",
    "philadelphia": "phoenix",
    "pittsburgh":   "thunderbirds",
    "salt-lake":    "shred",
    "san-diego":    "growlers",
    "seattle":      "cascades",
    "toronto":      "rush",
}

# Regex to match regular-season game slugs:
#   <away>-at-<home>-<M>-<D>-<YYYY>-<upload-date>-<duration>
# e.g. detroit-at-madison-6-27-2025-06-28-2025-02-08-00
#
# A second pattern handles the rare malformed slug where an extra number
# appears in the date: carolina-at-atlanta-5-3-17-2025-... (May 17 game)
_DATE_SUFFIX = r"-\d{2}-\d{2}-\d{4}-\d{2}-\d{2}-\d{2}$"  # upload-date + duration
_TEAMS       = r"^(?P<away>[a-z][a-z-]*)-at-(?P<home>[a-z][a-z-]*)"

GAME_SLUG_RE          = re.compile(_TEAMS + r"-(?P<month>\d{1,2})-(?P<day>\d{1,2})-(?P<year>\d{4})" + _DATE_SUFFIX)
GAME_SLUG_RE_MALFORMED = re.compile(_TEAMS + r"-(?P<month>\d{1,2})-\d{1,2}-(?P<day>\d{1,2})-(?P<year>\d{4})" + _DATE_SUFFIX)

# ─── PARSE SITEMAP ────────────────────────────────────────────────────────────

def parse_sitemap(sitemap_path):
    """
    Returns a list of dicts:
      { away_id, home_id, game_date (date), url }
    One entry per unique game (duplicates across team collections removed).
    """
    tree = ElementTree.parse(sitemap_path)
    root = tree.getroot()
    ns   = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

    seen_slugs = set()
    games      = []

    for loc_el in root.iter("{http://www.sitemaps.org/schemas/sitemap/0.9}loc"):
        url  = loc_el.text.strip()
        # Only consider /videos/ paths
        m = re.search(r"/videos/(.+)$", url)
        if not m:
            continue
        slug = m.group(1)

        if slug in seen_slugs:
            continue

        match = GAME_SLUG_RE.match(slug) or GAME_SLUG_RE_MALFORMED.match(slug)
        if not match:
            continue

        away_slug = match.group("away")
        home_slug = match.group("home")
        year      = int(match.group("year"))
        month     = int(match.group("month"))
        day       = int(match.group("day"))

        if year != YEAR_FILTER:
            continue

        away_id = CITY_SLUG_TO_TEAM.get(away_slug)
        home_id = CITY_SLUG_TO_TEAM.get(home_slug)
        if not away_id or not home_id:
            print(f"  WARNING: Unknown city slug in: {slug}")
            continue

        seen_slugs.add(slug)
        games.append({
            "away_id":   away_id,
            "home_id":   home_id,
            "game_date": date(year, month, day),
            "url":       url,
        })

    print(f"Parsed {len(games)} unique 2025 games from sitemap.")
    return games


# ─── LOAD GAMES FROM DATABASE ─────────────────────────────────────────────────

def load_games_from_db():
    url = (
        f"{SUPABASE_URL}/rest/v1/games"
        f"?Year=eq.{YEAR_FILTER}"
        f"&select=GameID,HomeTeamID,AwayTeamID,StartTimestamp"
        f"&order=StartTimestamp.asc"
    )
    req = urllib.request.Request(url, headers={
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    })
    with urllib.request.urlopen(req) as r:
        games = json.loads(r.read())
    print(f"Loaded {len(games)} games from database (Year={YEAR_FILTER})")
    return games


# ─── MATCH ────────────────────────────────────────────────────────────────────

def match_games(db_games, sitemap_games):
    """
    Builds a lookup: (away_id, home_id, game_date) → watchufa url
    Then annotates each db_game with its watchufa url (or None).

    DB timestamps are stored in UTC. UFA games played in the evening (ET/CT)
    often tip off at 7–8pm local = midnight+ UTC, so the UTC date can be one
    day ahead of the local game date used in the sitemap slug. We try both.
    """
    sitemap_lookup = {
        (g["away_id"], g["home_id"], g["game_date"]): g["url"]
        for g in sitemap_games
    }

    results = []
    matched = 0

    for g in db_games:
        ts        = datetime.fromisoformat(g["StartTimestamp"])
        utc_date  = ts.date()
        prev_date = utc_date - timedelta(days=1)

        key_utc  = (g["AwayTeamID"], g["HomeTeamID"], utc_date)
        key_prev = (g["AwayTeamID"], g["HomeTeamID"], prev_date)
        watch_url = sitemap_lookup.get(key_utc) or sitemap_lookup.get(key_prev)

        if watch_url:
            matched += 1

        results.append({
            "GameID":      g["GameID"],
            "Date":        ts.strftime("%Y-%m-%d"),
            "AwayTeamID":  g["AwayTeamID"],
            "HomeTeamID":  g["HomeTeamID"],
            "WatchUFAURL": watch_url or "",
            "Found":       watch_url is not None,
        })

    not_matched_in_db = matched
    extra_in_sitemap  = len(sitemap_games) - matched

    print(f"Matched {matched}/{len(db_games)} database games to a WatchUFA URL.")
    if extra_in_sitemap > 0:
        print(f"  ({extra_in_sitemap} sitemap games had no matching database entry — likely pre-season/extra content)")

    return results


# ─── OUTPUT: XLSX ─────────────────────────────────────────────────────────────

def write_xlsx(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WatchUFA Links"

    headers    = ["GameID", "Date", "Away", "Home", "WatchUFA URL"]
    col_widths = [22, 12, 10, 10, 80]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    found_fill   = PatternFill("solid", fgColor="C6EFCE")
    missing_fill = PatternFill("solid", fgColor="D9D9D9")
    data_font    = Font(name="Arial", size=10)
    link_font    = Font(name="Arial", size=10, color="0563C1", underline="single")

    for row_idx, r in enumerate(results, 2):
        row_fill = found_fill if r["Found"] else missing_fill
        values   = [r["GameID"], r["Date"], r["AwayTeamID"], r["HomeTeamID"], r["WatchUFAURL"]]
        for col, val in enumerate(values, 1):
            cell      = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            if col == 5 and val:
                cell.hyperlink = val
                cell.font      = link_font
            else:
                cell.font = data_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary sheet
    ws2     = wb.create_sheet("Summary")
    found   = sum(1 for r in results if r["Found"])
    missing = len(results) - found
    bold    = Font(bold=True, name="Arial", size=11)
    norm    = Font(name="Arial", size=11)
    for i, (label, val) in enumerate([
        ("Total games",        len(results)),
        ("✅ WatchUFA found",  found),
        ("❌ Not found",       missing),
    ], 1):
        ws2.cell(row=i, column=1, value=label).font = bold
        ws2.cell(row=i, column=2, value=val).font   = norm
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10

    wb.save(OUTPUT_XLSX)
    print(f"\nSpreadsheet saved: {OUTPUT_XLSX} ({len(results)} total rows)")


# ─── OUTPUT: SQL ──────────────────────────────────────────────────────────────

def write_sql(results):
    lines = [
        "-- WatchUFA URL updates",
        "-- REVIEW the spreadsheet before running this!",
        "",
        "-- Add the column if it doesn't exist:",
        'ALTER TABLE public.games ADD COLUMN IF NOT EXISTS "WatchUFAURL" text;',
        "",
        "BEGIN;",
        "",
    ]
    included = 0
    for r in results:
        if r["WatchUFAURL"]:
            safe_url = r["WatchUFAURL"].replace("'", "''")
            lines.append(
                f"UPDATE public.games SET \"WatchUFAURL\" = '{safe_url}'\n"
                f"  WHERE \"GameID\" = '{r['GameID']}';"
            )
            included += 1

    lines += ["", "COMMIT;"]
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"SQL saved: {OUTPUT_SQL} ({included} updates)")


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    sitemap_games = parse_sitemap(SITEMAP_PATH)
    db_games      = load_games_from_db()
    results       = match_games(db_games, sitemap_games)
    write_xlsx(results)
    write_sql(results)
    print("\nDone!")

if __name__ == "__main__":
    main()
