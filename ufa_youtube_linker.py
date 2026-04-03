#!/usr/bin/env python3
"""
UFA Game YouTube Link Finder
==============================
Searches YouTube for UFA game highlight videos and outputs a spreadsheet
with GameID → YouTube URL mappings, ready to import into your database.

SETUP:
  pip install google-api-python-client openpyxl pandas python-dotenv

.env file needs:
  YOUTUBE_API_KEY=...
  SUPABASE_URL=...
  SUPABASE_KEY=...

RESUMING after quota / interruption:
  Just re-run! The script automatically detects already-processed GameIDs
  from the existing spreadsheet and skips them. No need to set START_FROM.
"""

import os
import time
import json
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

# ─── CONFIGURATION ────────────────────────────────────────────────────────────

YOUTUBE_API_KEY = os.environ["YOUTUBE_API_KEY"]
SUPABASE_URL    = os.environ["SUPABASE_URL"]
SUPABASE_KEY    = os.environ["SUPABASE_KEY"]

INPUT_MODE     = "database"  # "database" or "csv"
YEAR_FILTER    = 2025
CSV_INPUT_PATH = "games.csv"  # only if INPUT_MODE = "csv"

OUTPUT_XLSX = "ufa_youtube_links.xlsx"
OUTPUT_SQL  = "ufa_youtube_update.sql"
PROGRESS_JSON = "ufa_youtube_progress.json"

# Set to True to force a clean restart from the first game of YEAR_FILTER.
# Safety behavior: clearing only happens when no checkpoint exists yet, so a
# quota-stopped rerun can still resume correctly.
RESTART_FROM_SCRATCH = True

SEARCH_DELAY_SECONDS = 1.0

# Official UFA YouTube channel ID
UFA_CHANNEL_ID = "UCzInURHrtSH7208Mf1HVqUA"

# ─── TEAM NAMES ───────────────────────────────────────────────────────────────
# First entry = used in search query
# All entries = checked against video title for confidence scoring

TEAM_NAMES = {
    "ATL":  ["Atlanta Hustle", "Hustle"],
    "AUS":  ["Austin Sol", "Sol"],
    "BAL":  ["Baltimore Whitewall", "Whitewall"],
    "BOS":  ["Boston Glory", "Glory"],
    "CAR":  ["Carolina Flyers", "Flyers"],
    "CHI":  ["Chicago Union", "Union"],
    "CIN":  ["Cincinnati Goats", "Goats"],
    "CLT":  ["Charlotte Express", "Express"],
    "COL":  ["Colorado Apex", "Apex", "Colorado Summit", "Summit"],  # rebranded from Summit
    "DAL":  ["Dallas Roughnecks", "Roughnecks"],
    "DC":   ["DC Breeze", "Breeze"],
    "DET":  ["Detroit Mechanix", "Mechanix"],
    "HOU":  ["Houston Havoc", "Havoc"],
    "IND":  ["Indianapolis AlleyCats", "AlleyCats", "Indy AlleyCats"],
    "LA":   ["Los Angeles Aviators", "Aviators"],
    "MAD":  ["Madison Radicals", "Radicals"],
    "MIA":  ["Miami Mutiny", "Mutiny"],
    "MIN":  ["Minnesota Wind Chill", "Wind Chill", "Windchill"],
    "MON":  ["Montreal Royal", "Royal"],
    "NY":   ["New York Empire", "Empire"],
    "OAK":  ["Oakland Spiders", "Spiders"],
    "OR":   ["Oregon Steel", "Steel"],
    "PHI":  ["Philadelphia Phoenix", "Phoenix"],
    "PIT":  ["Pittsburgh Thunderbirds", "Thunderbirds"],
    "POR":  ["Portland Nitro", "Nitro"],
    "RAL":  ["Raleigh Flyers", "Flyers"],
    "SD":   ["San Diego Growlers", "Growlers"],
    "SEA":  ["Seattle Cascades", "Cascades"],
    "SL":   ["Salt Lake Shred", "Shred"],
    "STL":  ["St. Louis Arch Rivals", "Arch Rivals"],
    "TOR":  ["Toronto Rush", "Rush"],
    "VAN":  ["Vancouver Riptide", "Riptide"],
    "VEG":  ["Vegas Bighorns", "Bighorns"],
    "WAS":  ["Washington DC Shadow", "Shadow", "Washington Shadow"],
}

def primary_name(team_id):
    names = TEAM_NAMES.get(team_id)
    return names[0] if names else team_id

def all_names(team_id):
    return TEAM_NAMES.get(team_id, [team_id])

# ─── DEPENDENCIES ─────────────────────────────────────────────────────────────

try:
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
except ImportError:
    raise ImportError("Run: pip install google-api-python-client")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("Run: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    raise ImportError("Run: pip install pandas")

# ─── LOAD GAMES ───────────────────────────────────────────────────────────────

def load_games_from_db():
    import urllib.request, json
    url = (
        f"{SUPABASE_URL}/rest/v1/games"
        f"?Year=eq.{YEAR_FILTER}"
        f"&select=GameID,HomeTeamID,AwayTeamID,StartTimestamp"
        f"&order=StartTimestamp.asc"
    )
    req = urllib.request.Request(url, headers={
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    })
    with urllib.request.urlopen(req) as r:
        games = json.loads(r.read())
    print(f"Loaded {len(games)} games from database (Year={YEAR_FILTER})")
    return games

def load_games_from_csv():
    df = pd.read_csv(CSV_INPUT_PATH, parse_dates=["StartTimestamp"])
    games = df[["GameID", "HomeTeamID", "AwayTeamID", "StartTimestamp"]].to_dict("records")
    print(f"Loaded {len(games)} games from {CSV_INPUT_PATH}")
    return games

def load_games():
    if INPUT_MODE == "database":
        return load_games_from_db()
    elif INPUT_MODE == "csv":
        return load_games_from_csv()
    else:
        raise ValueError(f"Unknown INPUT_MODE: {INPUT_MODE!r}")

# ─── LOAD EXISTING RESULTS ────────────────────────────────────────────────────

def load_existing_results():
    """
    Reads the existing spreadsheet and returns (results_list, done_game_ids).
    If no spreadsheet exists yet, returns ([], set()).
    This is how we avoid overwriting previous work.
    """
    if not os.path.exists(OUTPUT_XLSX):
        return [], set()

    try:
        wb  = openpyxl.load_workbook(OUTPUT_XLSX)
        ws  = wb["YouTube Links"]
        headers = [cell.value for cell in ws[1]]

        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # skip blank rows
                continue
            r = dict(zip(headers, row))
            results.append({
                "GameID":      r.get("GameID", ""),
                "Date":        r.get("Date", ""),
                "AwayTeamID":  r.get("Away", ""),
                "HomeTeamID":  r.get("Home", ""),
                "SearchQuery": r.get("Search Query Used", ""),
                "YouTubeURL":  r.get("YouTube URL", "") or "",
                "VideoTitle":  r.get("Video Title Found", "") or "",
                "Confidence":  r.get("Confidence", ""),
            })

        done_ids = {r["GameID"] for r in results}
        print(f"Found existing spreadsheet with {len(results)} results — will skip those GameIDs.")
        return results, done_ids

    except Exception as e:
        print(f"⚠️  Could not read existing spreadsheet ({e}). Starting fresh.")
        return [], set()

def clear_previous_outputs():
    removed = []
    failed = []
    for path in (OUTPUT_XLSX, OUTPUT_SQL, PROGRESS_JSON):
        if os.path.exists(path):
            try:
                os.remove(path)
                removed.append(path)
            except OSError as e:
                failed.append(f"{path} ({e})")
    if removed:
        print(f"🧹 Cleared previous run files: {', '.join(removed)}")
    if failed:
        print(f"⚠️  Could not clear: {', '.join(failed)}")
        print("   Close any app using these files (like Excel), then re-run if needed.")
    if not removed and not failed:
        print("🧹 No previous run files found to clear.")

def load_progress(total_games):
    """
    Returns the index of the next game to process.
    Falls back safely to 0 if checkpoint is missing/corrupt.
    """
    if not os.path.exists(PROGRESS_JSON):
        return 0
    try:
        with open(PROGRESS_JSON, "r", encoding="utf-8") as f:
            data = json.load(f)
        idx = int(data.get("next_index", 0))
        if idx < 0 or idx > total_games:
            print("⚠️  Progress checkpoint out of range, restarting from 0.")
            return 0
        print(f"↪️  Resuming from game #{idx + 1} of {total_games}.")
        return idx
    except Exception as e:
        print(f"⚠️  Could not read progress checkpoint ({e}). Restarting from 0.")
        return 0

def save_progress(next_index, total_games):
    data = {
        "year": YEAR_FILTER,
        "input_mode": INPUT_MODE,
        "next_index": next_index,
        "total_games": total_games,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    with open(PROGRESS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def clear_progress_file():
    if os.path.exists(PROGRESS_JSON):
        os.remove(PROGRESS_JSON)
        print("🗑️  Progress checkpoint cleared (run completed).")

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def parse_timestamp(ts):
    if isinstance(ts, str):
        return datetime.fromisoformat(ts)
    return ts

def format_date(ts):
    """Returns date like 'June 7, 2025' (no leading zero) — matches YouTube title format."""
    return ts.strftime("%B %d, %Y").replace(" 0", " ")

# ─── SEARCH QUERY ─────────────────────────────────────────────────────────────

def build_search_query(game):
    home     = primary_name(game["HomeTeamID"])
    away     = primary_name(game["AwayTeamID"])
    ts       = parse_timestamp(game["StartTimestamp"])
    date_str = format_date(ts)
    return f"{away} at {home} highlights {date_str} UFA"

# ─── CONFIDENCE SCORING ───────────────────────────────────────────────────────

def score_match(title, game):
    """
    HIGH   → correct teams + correct date in title
    MEDIUM → correct teams but date missing/wrong
    LOW    → only one team found, or no highlight keywords
    WRONG  → neither team found (excluded from output)
    """
    t  = title.lower()
    ts = parse_timestamp(game["StartTimestamp"])

    home_names = [n.lower() for n in all_names(game["HomeTeamID"])]
    away_names = [n.lower() for n in all_names(game["AwayTeamID"])]

    home_match = any(n in t for n in home_names)
    away_match = any(n in t for n in away_names)

    date_variants = [
        format_date(ts),
        ts.strftime("%B %d, %Y"),
        ts.strftime("%b. %d, %Y").replace(" 0", " "),
        ts.strftime("%b %d, %Y").replace(" 0", " "),
        # Uppercase month variants (e.g. "AUG 9, 2025" seen in playoff titles)
        format_date(ts).upper(),
        ts.strftime("%b %d, %Y").replace(" 0", " ").upper(),
    ]
    date_match = any(d.lower() in t for d in date_variants)

    is_highlights = any(w in t for w in [
        "highlight", "full game", "full match",
        "playoffs", "championship", "semifinal", "division",
    ])

    if not is_highlights:
        return "LOW"
    if home_match and away_match and date_match:
        return "HIGH"
    if home_match and away_match:
        return "MEDIUM"
    if home_match or away_match:
        return "LOW"
    return "WRONG"

# ─── YOUTUBE SEARCH ───────────────────────────────────────────────────────────

def search_youtube(youtube, query, game):
    try:
        response = youtube.search().list(
            q=query,
            part="snippet",
            type="video",
            maxResults=5,
            order="relevance",
            channelId=UFA_CHANNEL_ID,
        ).execute()
    except HttpError as e:
        if e.resp.status == 403:
            raise
        return None, None, "ERROR"

    items = response.get("items", [])
    if not items:
        return None, None, "NOT FOUND"

    rank   = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "WRONG": 3}
    scored = sorted(
        [(score_match(item["snippet"]["title"], game),
          item["id"]["videoId"],
          item["snippet"]["title"])
         for item in items],
        key=lambda x: rank.get(x[0], 99),
    )

    best_conf, best_id, best_title = scored[0]
    if best_conf == "WRONG":
        return None, None, "NOT FOUND"

    return f"https://www.youtube.com/watch?v={best_id}", best_title, best_conf

# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    youtube = build("youtube", "v3", developerKey=YOUTUBE_API_KEY)

    force_fresh_start = RESTART_FROM_SCRATCH and not os.path.exists(PROGRESS_JSON)
    if force_fresh_start:
        clear_previous_outputs()

    games   = load_games()
    total   = len(games)

    if force_fresh_start:
        # Ignore previous spreadsheet contents; restart from first game in year.
        results = []
        start_idx = 0
        print(f"🔁 Forced fresh start: beginning at game 1 of {total}.")
    else:
        # Load previously saved rows for this run's resume behavior
        results, _ = load_existing_results()
        start_idx = load_progress(total)

    if start_idx >= total:
        print("All games already processed! Check your spreadsheet.")
        clear_progress_file()
        return

    try:
        for idx in range(start_idx, total):
            game    = games[idx]
            overall = idx + 1
            query     = build_search_query(game)
            print(f"[{overall}/{total}] Searching: {query}")

            try:
                url, title, confidence = search_youtube(youtube, query, game)
            except HttpError as e:
                if e.resp.status == 403:
                    print(f"\n🚫 Quota exceeded at game {overall}/{total}.")
                    print(f"   Swap in a new API key and just re-run — progress is saved!")
                    _save(results, idx, total)
                    return
                raise

            ts = parse_timestamp(game["StartTimestamp"])
            results.append({
                "GameID":      game["GameID"],
                "Date":        ts.strftime("%Y-%m-%d"),
                "AwayTeamID":  game["AwayTeamID"],
                "HomeTeamID":  game["HomeTeamID"],
                "SearchQuery": query,
                "YouTubeURL":  url or "",
                "VideoTitle":  title or "",
                "Confidence":  confidence,
            })

            icons = {"HIGH": "✅", "MEDIUM": "🟡", "LOW": "🔶", "NOT FOUND": "❌", "ERROR": "⚠️"}
            if url:
                print(f"   {icons.get(confidence,'❓')} {confidence}: {title}")
                print(f"      {url}")
            else:
                print(f"   {icons.get(confidence,'❓')} {confidence}")

            # Save after each game so resume is exact, even after abrupt stops.
            _save(results, idx + 1, total)
            time.sleep(SEARCH_DELAY_SECONDS)

    except KeyboardInterrupt:
        print(f"\n⚠️  Interrupted. Progress saved — just re-run to continue!")
        next_index = len(results)
        _save(results, next_index, total)
        return

    _save(results, total, total)
    clear_progress_file()
    print(f"\n✅ All done!")

# ─── SAVE (always appends/overwrites with full combined results) ───────────────

def _save(results, next_index, total_games):
    if not results:
        print("No results to save.")
        save_progress(next_index, total_games)
        return
    write_xlsx(results)
    write_sql(results)
    save_progress(next_index, total_games)

# ─── OUTPUT: XLSX ─────────────────────────────────────────────────────────────

def write_xlsx(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube Links"

    headers    = ["GameID", "Date", "Away", "Home", "YouTube URL", "Video Title Found", "Confidence", "Search Query Used"]
    col_widths = [22, 12, 10, 10, 52, 65, 12, 65]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    conf_colors = {
        "HIGH":      "C6EFCE",
        "MEDIUM":    "FFEB9C",
        "LOW":       "FFD966",
        "NOT FOUND": "D9D9D9",
        "WRONG":     "FFC7CE",
        "ERROR":     "FFC7CE",
    }
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for row_idx, r in enumerate(results, 2):
        conf     = r["Confidence"]
        row_fill = PatternFill("solid", fgColor=conf_colors.get(conf, "FFFFFF"))
        values   = [
            r["GameID"], r["Date"], r["AwayTeamID"], r["HomeTeamID"],
            r["YouTubeURL"], r["VideoTitle"], conf, r["SearchQuery"],
        ]
        for col, val in enumerate(values, 1):
            cell      = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            if col == 5 and val:
                cell.hyperlink = val
                cell.font = link_font
            else:
                cell.font = data_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary sheet
    ws2    = wb.create_sheet("Summary")
    counts = {k: sum(1 for r in results if r["Confidence"] == k)
              for k in ["HIGH", "MEDIUM", "LOW", "WRONG", "NOT FOUND", "ERROR"]}
    bold   = Font(bold=True, name="Arial", size=11)
    norm   = Font(name="Arial", size=11)
    for i, (label, val) in enumerate([
        ("Total processed",      len(results)),
        ("✅ HIGH confidence",   counts["HIGH"]),
        ("🟡 MEDIUM confidence", counts["MEDIUM"]),
        ("🔶 LOW confidence",    counts["LOW"]),
        ("❌ WRONG / Not found", counts["WRONG"] + counts["NOT FOUND"]),
        ("⚠️  Errors",           counts["ERROR"]),
    ], 1):
        ws2.cell(row=i, column=1, value=label).font = bold
        ws2.cell(row=i, column=2, value=val).font   = norm
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10

    wb.save(OUTPUT_XLSX)
    print(f"\n📊 Spreadsheet saved: {OUTPUT_XLSX} ({len(results)} total rows)")

# ─── OUTPUT: SQL ──────────────────────────────────────────────────────────────

def write_sql(results):
    lines = [
        "-- UFA YouTube URL updates",
        "-- REVIEW the spreadsheet before running this!",
        "-- Includes HIGH and MEDIUM confidence matches.",
        "",
        "-- Add the column if it doesn't exist:",
        'ALTER TABLE public.games ADD COLUMN IF NOT EXISTS "YouTubeURL" text;',
        "",
        "BEGIN;",
        "",
    ]
    included = 0
    for r in results:
        if r["YouTubeURL"] and r["Confidence"] in ("HIGH", "MEDIUM"):
            safe_url   = r["YouTubeURL"].replace("'", "''")
            safe_title = r["VideoTitle"].replace("'", "''")
            lines.append(
                f"UPDATE public.games SET \"YouTubeURL\" = '{safe_url}'"
                f"  -- {safe_title}\n"
                f"  WHERE \"GameID\" = '{r['GameID']}';"
            )
            included += 1
    lines += ["", "COMMIT;"]
    with open(OUTPUT_SQL, "w") as f:
        f.write("\n".join(lines))
    print(f"💾 SQL file saved: {OUTPUT_SQL} ({included} updates)")

if __name__ == "__main__":
    main()